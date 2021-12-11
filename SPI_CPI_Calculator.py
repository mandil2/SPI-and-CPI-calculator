#Vikash Mandil- 1901ME68

# this function will create overall sheet
def overall_sheet():

    with open('names-roll.csv', 'r') as file:

        reader=csv.reader(file)
        for row in reader:
            if row[0] != "Roll":

                sheet_path = "output\\"+'%s.xlsx' %row[0]
                wb = load_workbook(sheet_path)
                sheet1 = wb["Sheet"]

                row_1 = ["Roll No.", row[0]]        #Roll no. row
                row_2 = ["Name of Student", row[1]] #name row
                row_3 = ["Discipline", row[0][4:6]] #branch row
    
                row_4 = []                          #semester no. row
                row_4.append("Semester No.")

                row_5 = []                          # present semester credits row
                row_5.append("Semester wise Credits taken")

                row_6 = []                          #present semester spi row
                row_6.append("SPI")

                row_7 = []                          # total credits upto present semester row
                row_7.append("Total Credit taken")

                row_8 = []                          # cpi upto present semester row
                row_8.append("CPI")
               
                numeric_grade = {"AA": 10, "AA*": 10,"AB": 9, "AB*": 9,"BB": 8, "BB*": 8,
                        "BC": 7, "CC": 6, "CD": 5, "DD": 4, "DD*": 4, "F": 0,"F*":0, "I": 0}

                total_score = 0
                total_credit = 0
                
                for i in range(1, len(wb.sheetnames)+1): 
                    
                    sheet_name = "Sem" + str(i)
                    
                    if sheet_name in wb.sheetnames:
                        row_4.append(i)
                        sem_sheet = wb[sheet_name]

                    else:
                        continue
                    
                    sem_credit = 0
                    sem_score = 0

                    for j in range(2, sem_sheet.max_row + 1):

                        cur_credit = sem_sheet.cell(row = j, column = 5).value
                        cur_grade = numeric_grade[sem_sheet.cell(row = j, column = 7).value]
                        
                        sem_credit += cur_credit

                        sem_score += (cur_grade * cur_credit)
                    
                    total_credit += sem_credit
                    total_score += sem_score
                    
                    spi = sem_score / sem_credit
                    spi = round(spi, 2)

                    cpi = total_score / total_credit
                    cpi = round(cpi, 2)
                    
                    row_5.append(sem_credit)
                    row_6.append(spi)
                    row_7.append(total_credit)
                    row_8.append(cpi)
                        
                sheet1.append(row_1)
                sheet1.append(row_2)
                sheet1.append(row_3)
                sheet1.append(row_4)
                sheet1.append(row_5)
                sheet1.append(row_6)
                sheet1.append(row_7)
                sheet1.append(row_8)
                
                sheet1.title = "Overall"
                
                wb.save(sheet_path)

    return       

def generate_marksheet():
    # opening the grades.csv file in read mode
    with open('grades.csv','r') as file:
        reader=csv.reader(file)
        header=['Sl No.','Subject No.','Subject Name','L-T-P','Credit','Subject Type','Grade']
        for data in reader:
            if data[0]!="Roll":
                sheet_path="output\\"+'%s.xlsx' %data[0]
            
                # if the excel file with the name already exits
                if os.path.exists(sheet_path):
                    wb=load_workbook(sheet_path)

                    sheet_name='Sem'+str(data[1])

                    # if the sheet with the name already exists
                    if sheet_name in wb.sheetnames:            
                        sheet1=wb[sheet_name]
                        sl_no=sheet1.max_row
                        
                        with open('subjects_master.csv','r') as file:
                            reader=csv.reader(file)
                            for row in reader:
                                if row[0]==data[2]:
                                    lst=[sl_no,row[0],row[1],row[2],int(row[3]),data[-1],(data[-2].strip())]
                                    break
                        sheet1.append(lst)

                    # if the sheet with the name does not  exist
                    else:
                        sheet1=wb.create_sheet(sheet_name)
                        sheet1.append(header)
                        sl_no=sheet1.max_row

                        with open('subjects_master.csv','r') as file:
                            reader=csv.reader(file)
                            for row in reader:
                                if row[0]==data[2]:
                                    lst=[sl_no,row[0],row[1],row[2],int(row[3]),data[-1],(data[-2].strip())]
                                    break
                        sheet1.append(lst)

                    wb.save(sheet_path)

                # if the excel file does not already exists
                else :

                    wb = Workbook()
                    sheet_name='Sem'+str(data[1])

                    # if the sheet with the name already exists
                    if sheet_name in wb.sheetnames:
                        sheet1=wb[sheet_name]

                        sl_no=sheet1.max_row
                        with open('subjects_master.csv','r') as file:

                            reader=csv.reader(file)
                            for row in reader:
                                if row[0]==data[2]:
                                    lst=[sl_no,row[0],row[1],row[2],int(row[3]),data[-1],(data[-2].strip())]
                                    break
                        sheet1.append(lst)

                    # if the sheet with the name does not  exist

                    else:

                        sheet1=wb.create_sheet(sheet_name)
                        sheet1.append(header)
                        sl_no= sheet1.max_row


                        with open('subjects_master.csv','r') as file:
                            reader=csv.reader(file)

                            for row in reader:

                                if row[0]==data[2]:

                                    lst=[sl_no,row[0],row[1],row[2],int(row[3]),data[-1],(data[-2].strip())]

                                    break
                        sheet1.append(lst)
                    wb.save(sheet_path)
    
    # calling the overall sheet function
    overall_sheet()
    return




# main function
import os
import csv
from openpyxl import Workbook
from openpyxl import load_workbook

os.system("cls")

# create the output folder if it doesn't exist
if not os.path.exists(r"output\\"):
    os.mkdir('.\\output')

# calling the generate_marksheet function
generate_marksheet()

            




